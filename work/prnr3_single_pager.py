#!/usr/bin/env python3
import datetime as dt
import html
import json
import math
import statistics
import subprocess
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

import morning_summary as ms


OUT_DIR = Path("outputs")
SRC_DIR = OUT_DIR / "prnr3_single_pager_sources"
SHEET_PATH = SRC_DIR / "planilha_1t26.xlsx"
HTML_PATH = OUT_DIR / "prnr3-single-pager.html"
PDF_PATH = OUT_DIR / "prnr3-single-pager.pdf"
SYMBOL = "PRNR3.SA"
TICKER = "PRNR3"

SOURCE_LINKS = {
    "release_1t26": "https://api.mziq.com/mzfilemanager/v2/d/6aff4303-d559-4771-901f-8de7d8a88ef5/2f6ee62a-43b3-1731-fc81-ce17b0f1c94b?origin=2",
    "presentation_1t26": "https://api.mziq.com/mzfilemanager/v2/d/6aff4303-d559-4771-901f-8de7d8a88ef5/1a94c8e4-bc97-99d2-591b-a8397570ddf2?origin=2",
    "financial_sheet_1t26": "https://api.mziq.com/mzfilemanager/v2/d/6aff4303-d559-4771-901f-8de7d8a88ef5/39927c29-49c5-61d8-d7ee-14a916279237?origin=2",
    "analyst_coverage": "https://ri.priner.com.br/informacoes-aos-investidores/cobertura-de-analistas/",
    "yahoo": "https://finance.yahoo.com/quote/PRNR3.SA",
}

PREVIEW_2T26 = {
    "date": "06/07/2026",
    "gross_revenue_2t26": 515.0,
    "gross_revenue_1t26": 471.3,
    "gross_revenue_2t25": 394.4,
    "employees_end_2t26": 7990,
    "employees_end_1t26": 7741,
    "employees_end_2t25": 6602,
    "employees_avg_2t26": 8010,
    "employees_avg_1t26": 7818,
    "employees_avg_2t25": 6753,
    "new_contracts": 966.4,
    "open_positions": 957,
    "tp_adjustment": 0.08,
}


def main():
    OUT_DIR.mkdir(exist_ok=True)
    data = build_single_pager_data()
    html_content = render_html(data)
    HTML_PATH.write_text(html_content, encoding="utf-8")
    ok, msg = ms.make_chromium_pdf(HTML_PATH, PDF_PATH)
    if not ok:
        ok, msg = ms.make_pdf(HTML_PATH, PDF_PATH)
    if not ok:
        raise RuntimeError(f"PDF nao gerado: {msg}")
    print(f"HTML: {HTML_PATH}")
    print(f"PDF: {PDF_PATH}")


def build_single_pager_data():
    history = fetch_yahoo_history(SYMBOL, "1y", "1d")
    quote = fetch_yahoo_quote(SYMBOL)
    price = num(quote.get("regularMarketPrice")) or latest_close(history)
    previous_close = num(quote.get("regularMarketPreviousClose"))
    if previous_close is None and len(history) >= 2:
        previous_close = history[-2]["close"]
    market_cap = num(quote.get("marketCap"))
    if market_cap is None and price is not None:
        market_cap = price * 56_720_000
    day_change = None
    if price is not None and previous_close:
        day_change = (price / previous_close - 1) * 100

    financials = load_financials()
    consensus_targets = [
        {"house": "BTG Pactual", "target": 28.00, "rating": "Compra", "date": "Fev/26", "call_date": "2026-02-01"},
        {"house": "Itau BBA", "target": 30.00, "rating": "Outperform", "date": "Mar/26", "call_date": "2026-03-01"},
        {"house": "XP", "target": 22.10, "rating": "Compra", "date": "Nov/25", "call_date": "2025-11-01"},
        {"house": "Empiricus", "target": 20.00, "rating": "Compra", "date": "Out/24", "call_date": "2024-10-01"},
    ]
    consensus_value = statistics.mean(item["target"] for item in consensus_targets)

    ms.CANDIDATE_SYMBOLS[TICKER] = SYMBOL
    ms.CANDIDATE_IS_ETF[TICKER] = False
    ms.CANDIDATE_REQUIRED_UPSIDE[TICKER] = 0.33
    ms.CANDIDATE_BUY_IN_DISCOUNT[TICKER] = 0.11
    item = {
        "ticker": TICKER,
        "symbol": SYMBOL,
        "sector": "Servicos industriais / infraestrutura",
        "quality": 62,
        "ai": 56,
        "cyclical": True,
        "catalyst": "backlog recorde, normalizacao de mineracao, disciplina de SG&A e maturacao de contratos",
        "risk": "chuvas/mineracao, alavancagem, execucao de M&A e reoneracao da folha ate 2027",
    }
    multiples = ms.fetch_candidate_multiples(TICKER, SYMBOL, "")
    price_context = ms.fetch_candidate_price_context(SYMBOL)
    model = ms.candidate_modeled_target_snapshot(item, price, consensus_value, multiples, price_context)
    base_our_tp = model.get("target_value") or consensus_value
    our_tp = base_our_tp * (1 + PREVIEW_2T26["tp_adjustment"])
    buy_in = ms.candidate_buy_in_value(TICKER, price, our_tp, price_context, multiples)
    summary_snapshot = ms.latest_summary_candidate_snapshot(TICKER)
    if summary_snapshot:
        snapshot_price = summary_snapshot.get("price")
        snapshot_our_tp = summary_snapshot.get("our_tp")
        snapshot_buy_in = summary_snapshot.get("buy_in")
        snapshot_consensus = summary_snapshot.get("consensus")
        if snapshot_price is not None:
            price = snapshot_price
        if snapshot_our_tp is not None:
            if our_tp and model.get("methods"):
                factor = snapshot_our_tp / our_tp
                model["methods"] = {
                    name: (value * factor if value else value)
                    for name, value in model.get("methods", {}).items()
                }
            our_tp = snapshot_our_tp
            base_our_tp = our_tp / (1 + PREVIEW_2T26["tp_adjustment"])
        if snapshot_buy_in is not None:
            buy_in = snapshot_buy_in
        if snapshot_consensus is not None:
            consensus_value = snapshot_consensus

    return {
        "as_of": dt.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "price": price,
        "day_change": day_change,
        "market_cap": market_cap,
        "history": history,
        "financials": financials,
        "multiples": multiples,
        "consensus_targets": consensus_targets,
        "consensus_value": consensus_value,
        "our_tp": our_tp,
        "model_methods": model.get("methods", {}),
        "base_our_tp": base_our_tp,
        "preview": build_preview_metrics(),
        "buy_in": buy_in,
        "upside": pct_change(price, our_tp),
        "consensus_upside": pct_change(price, consensus_value),
        "buy_in_distance": pct_change(buy_in, price) if buy_in else None,
        "source_links": SOURCE_LINKS,
    }


def build_preview_metrics():
    preview = dict(PREVIEW_2T26)
    preview["gross_revenue_qoq"] = pct_change(
        preview["gross_revenue_1t26"],
        preview["gross_revenue_2t26"],
    )
    preview["gross_revenue_yoy"] = pct_change(
        preview["gross_revenue_2t25"],
        preview["gross_revenue_2t26"],
    )
    preview["employees_end_yoy"] = pct_change(
        preview["employees_end_2t25"],
        preview["employees_end_2t26"],
    )
    preview["employees_avg_yoy"] = pct_change(
        preview["employees_avg_2t25"],
        preview["employees_avg_2t26"],
    )
    preview["revenue_per_avg_employee_2t26"] = (
        preview["gross_revenue_2t26"] / preview["employees_avg_2t26"]
    )
    preview["revenue_per_avg_employee_1t26"] = (
        preview["gross_revenue_1t26"] / preview["employees_avg_1t26"]
    )
    preview["revenue_per_avg_employee_yoy"] = pct_change(
        preview["gross_revenue_2t25"] / preview["employees_avg_2t25"],
        preview["revenue_per_avg_employee_2t26"],
    )
    preview["book_to_bill_quarter"] = (
        preview["new_contracts"] / preview["gross_revenue_2t26"]
    )
    return preview


def fetch_yahoo_quote(symbol):
    url = "https://query1.finance.yahoo.com/v7/finance/quote?symbols=" + urllib.parse.quote(symbol)
    try:
        with urllib.request.urlopen(request(url), timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8"))
        rows = payload.get("quoteResponse", {}).get("result", [])
        return rows[0] if rows else {}
    except Exception:
        return {}


def fetch_yahoo_history(symbol, range_value, interval):
    url = (
        "https://query1.finance.yahoo.com/v8/finance/chart/"
        f"{urllib.parse.quote(symbol)}?range={urllib.parse.quote(range_value)}&interval={urllib.parse.quote(interval)}"
    )
    try:
        with urllib.request.urlopen(request(url), timeout=15) as response:
            payload = json.loads(response.read().decode("utf-8"))
        result = payload["chart"]["result"][0]
        timestamps = result.get("timestamp", [])
        quote = result.get("indicators", {}).get("quote", [{}])[0]
        closes = quote.get("close", [])
        points = []
        for ts, close in zip(timestamps, closes):
            if close is None or close <= 0:
                continue
            points.append({"date": dt.datetime.fromtimestamp(ts).date(), "close": float(close)})
        return points
    except Exception:
        return []


def request(url):
    return urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})


def load_financials():
    wb = load_workbook(SHEET_PATH, data_only=True, read_only=True)
    chart = wb["Dados para gráficos"]
    labels = row_values(chart, 4)[1:]
    rows = {
        "gross_revenue": row_values(chart, 5)[1:],
        "net_revenue": row_values(chart, 9)[1:],
        "gross_profit": row_values(chart, 13)[1:],
        "gross_margin": row_values(chart, 15)[1:],
        "ebitda_adj": row_values(chart, 26)[1:],
        "ebitda_margin_adj": row_values(chart, 28)[1:],
        "roic": row_values(chart, 31)[1:],
        "net_debt_ebitda": row_values(chart, 41)[1:],
        "revenue_per_capita": row_values(chart, 36)[1:],
        "gross_profit_per_capita": row_values(chart, 37)[1:],
    }
    latest = {key: values[-1] for key, values in rows.items() if values}
    prior_q = {key: values[-2] for key, values in rows.items() if len(values) >= 2}
    prior_y = {key: values[-5] for key, values in rows.items() if len(values) >= 5}
    ltm = {
        "net_revenue": sum(rows["net_revenue"][-4:]),
        "ebitda_adj": sum(rows["ebitda_adj"][-4:]),
        "gross_profit": sum(rows["gross_profit"][-4:]),
    }
    return {
        "labels": labels,
        "rows": rows,
        "latest": latest,
        "prior_q": prior_q,
        "prior_y": prior_y,
        "ltm": ltm,
    }


def row_values(ws, row_number):
    row = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))
    return [value for value in row if value is not None]


def render_html(data):
    f = data["financials"]
    latest = f["latest"]
    prior_q = f["prior_q"]
    prior_y = f["prior_y"]
    ltm = f["ltm"]
    p = data["preview"]
    history = data["history"]
    price_chart = sparkline_svg(history, 590, 155, data["consensus_targets"], show_legend=False)
    method_values = [value for value in data["model_methods"].values() if value]
    max_method = max(method_values) if method_values else None
    min_method = min(method_values) if method_values else None
    method_rows = "".join(
        valuation_method_row(name, value, max_method, min_method)
        for name, value in data["model_methods"].items()
        if value
    )
    consensus_rows = "".join(
        f"<tr><td>{html.escape(item['house'])}</td><td>{money(item['target'])}</td><td>{html.escape(item['rating'])}</td><td>{html.escape(item['date'])}</td></tr>"
        for item in data["consensus_targets"]
    )
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>PRNR3 Single Pager</title>
  <style>
    @page {{ size: A4 landscape; margin: 8mm; }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, Helvetica, sans-serif; color: #101828; background: #fff; }}
    .page {{ width: 281mm; min-height: 194mm; padding: 0; }}
    .top {{ display: grid; grid-template-columns: 1fr 1.25fr; gap: 7mm; align-items: end; border-bottom: 2px solid #d9e2ec; padding-bottom: 4mm; }}
    .eyebrow {{ color: #667085; font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: .7px; }}
    h1 {{ margin: 1mm 0 0; font-size: 34px; line-height: 1; letter-spacing: 0; }}
    .subtitle {{ color: #475467; font-size: 12px; margin-top: 2mm; }}
    .kpis {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 2mm; }}
    .kpi {{ border: 1px solid #d9e2ec; background: #f8fafc; padding: 2.2mm 2.5mm; min-height: 16mm; }}
    .kpi-label {{ font-size: 8px; color: #667085; text-transform: uppercase; font-weight: 850; letter-spacing: .4px; }}
    .kpi-value {{ font-size: 17px; font-weight: 950; margin-top: 1mm; white-space: nowrap; }}
    .kpi-sub {{ font-size: 8.3px; color: #667085; margin-top: .6mm; }}
    .pos {{ color: #067647; }} .neg {{ color: #b42318; }} .blue {{ color: #175cd3; }} .purple {{ color: #7c3aed; }}
    .valuation-high {{ color: #067647; font-weight: 850; }}
    .valuation-low {{ color: #b42318; font-weight: 850; }}
    .our-tp-row td {{ color: #175cd3; font-weight: 950; }}
    .grid {{ display: grid; grid-template-columns: 1.18fr .82fr; gap: 5mm; margin-top: 3.5mm; }}
    .left, .right {{ display: grid; gap: 2.6mm; }}
    .box {{ border: 1px solid #d9e2ec; padding: 2.6mm; background: #fff; break-inside: avoid; }}
    .compact-box {{ padding-top: 2mm; padding-bottom: 2mm; }}
    .box-title {{ font-size: 9.6px; color: #475467; font-weight: 900; text-transform: uppercase; letter-spacing: .6px; margin-bottom: 1.6mm; }}
    .chart-title-row {{ display: flex; align-items: center; justify-content: space-between; gap: 4mm; margin-bottom: 1.6mm; }}
    .chart-title-row .box-title {{ margin-bottom: 0; }}
    .call-legend {{ display: flex; align-items: center; gap: 4mm; color: #667085; font-size: 9px; font-weight: 500; text-transform: none; letter-spacing: 0; }}
    .call-legend span {{ display: inline-flex; align-items: center; gap: 1.4mm; }}
    .dot {{ width: 8px; height: 8px; border-radius: 50%; display: inline-block; }}
    .dot.buy {{ background: #16a34a; }}
    .dot.hold {{ background: #175cd3; }}
    .dot.sell {{ background: #111827; }}
    .chart-wrap {{ height: 45mm; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th {{ text-align: left; font-size: 7.8px; color: #667085; text-transform: uppercase; letter-spacing: .35px; background: #f2f5f8; padding: 4px 5px; }}
    td {{ font-size: 9.4px; padding: 4px 5px; border-top: 1px solid #edf1f5; vertical-align: top; }}
    .metric-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 1.6mm; }}
    .metric {{ background: #f8fafc; border: 1px solid #e7ebf0; padding: 1.8mm; min-height: 12mm; }}
    .metric b {{ display: block; font-size: 12px; margin-top: .7mm; }}
    .metric span {{ color: #667085; font-size: 7.8px; font-weight: 800; text-transform: uppercase; }}
    .bullets {{ margin: 0; padding-left: 4mm; }}
    .bullets li {{ font-size: 9.0px; line-height: 1.22; margin-bottom: 1.25mm; text-align: justify; text-justify: inter-word; }}
    .compact-bullets li {{ margin-bottom: .8mm; }}
    .tag {{ display: inline-block; padding: 1px 4px; border-radius: 3px; background: #eef4ff; color: #175cd3; font-size: 7px; font-weight: 900; text-transform: uppercase; }}
    .footer {{ position: absolute; bottom: 7mm; left: 8mm; right: 8mm; display: flex; justify-content: space-between; color: #98a2b3; font-size: 7.8px; border-top: 1px solid #edf1f5; padding-top: 2mm; }}
    a {{ color: inherit; text-decoration: none; }}
    .note {{ color: #667085; font-size: 8px; line-height: 1.2; margin-top: 1.2mm; text-align: justify; text-justify: inter-word; }}
  </style>
</head>
<body>
<section class="page">
  <div class="top">
    <div>
      <div class="eyebrow">Single Pager | B3 Equity Model</div>
      <h1>PRNR3 | Priner</h1>
      <div class="subtitle">Servicos industriais, infraestrutura, montagem e operacoes minerarias | Atualizado em {html.escape(data['as_of'])}</div>
    </div>
    <div class="kpis">
      {price_kpi(brl_plain(data["price"]), data["day_change"])}
      {kpi("Our TP", money(data["our_tp"]), signed_pct(data["upside"]) + " upside", "blue")}
      {kpi("Consensus", money(data["consensus_value"]), f"{len(data['consensus_targets'])} casas | {signed_pct(data['consensus_upside'])}", "purple")}
      {kpi("Buy-in", money(data["buy_in"]), signed_pct(data["buy_in_distance"]) + " vs preco")}
      {kpi("Mkt Cap", brl_big(data["market_cap"]), "Yahoo Finance")}
    </div>
  </div>

  <div class="grid">
    <div class="left">
      <div class="box">
        <div class="chart-title-row">
          <div class="box-title">Stock performance | ultimos 12 meses + analyst calls</div>
          <div class="call-legend">
            <span><i class="dot buy"></i>Buy</span>
            <span><i class="dot hold"></i>Hold</span>
            <span><i class="dot sell"></i>Sell</span>
          </div>
        </div>
        <div class="chart-wrap">{price_chart}</div>
        <div class="metric-grid">
          {metric("12m return", signed_pct(history_return(history)))}
          {metric("52w low/high", f"{brl_plain(history_low(history))} / {brl_plain(history_high(history))}")}
          {metric("Vol. anualizada", signed_pct(history_volatility(history), plus=False))}
          {metric("Ultimo fechamento", brl_plain(latest_close(history)))}
        </div>
      </div>

      <div class="box">
        <div class="box-title">Financial snapshot | previa operacional 2T26</div>
        <div class="metric-grid">
          {metric("ROB 2T26", brl_mm(p['gross_revenue_2t26']), f"QoQ {signed_pct(p['gross_revenue_qoq'])}")}
          {metric("ROB YoY", signed_pct(p['gross_revenue_yoy']), "vs 2T25")}
          {metric("Novos contratos", brl_mm(p['new_contracts']), f"{p['book_to_bill_quarter']:.2f}x ROB 2T26")}
          {metric("Vagas abertas", f"{p['open_positions']:,}".replace(",", "."), "preencher no 3T26")}
          {metric("Colab. final", f"{p['employees_end_2t26']:,}".replace(",", "."), f"YoY {signed_pct(p['employees_end_yoy'])}")}
          {metric("Colab. medio", f"{p['employees_avg_2t26']:,}".replace(",", "."), f"YoY {signed_pct(p['employees_avg_yoy'])}")}
          {metric("ROB/colab. medio", brl_mm(p['revenue_per_avg_employee_2t26']), f"YoY {signed_pct(p['revenue_per_avg_employee_yoy'])}")}
          {metric("EV/EBITDA", priner_ev_ebitda_display(data))}
        </div>
      </div>

      <div class="box compact-box">
        <div class="box-title">AI thesis & risk</div>
        <ul class="bullets compact-bullets">
          <li><b class="pos">Tese:</b> previa 2T26 confirmou aceleracao: ROB +30,6% YoY, R$ 966,4 mi em novos contratos e 957 vagas abertas sustentam crescimento futuro.</li>
          <li><b class="neg">Risco:</b> capital intensivo, alavancagem, clima/mineracao, reoneracao da folha ate 2027 e execucao de M&A podem atrasar ROIC.</li>
          <li><b>Trigger:</b> resultado completo do 2T26 precisa confirmar margem/EBITDA, capital de giro e conversao de caixa para validar novo TP.</li>
        </ul>
      </div>
    </div>

    <div class="right">
      <div class="box">
        <div class="box-title">Valuation model | nossa metodologia</div>
        <table>
          <thead><tr><th>Driver</th><th>Valor</th></tr></thead>
          <tbody>
            {method_rows}
            <tr><td>Previa operacional 2T26</td><td>+{data['preview']['tp_adjustment'] * 100:.0f}% ajuste preliminar</td></tr>
            <tr><td>TP antes da previa</td><td>{money(data['base_our_tp'])}</td></tr>
            <tr class="our-tp-row"><td>Our TP medio</td><td>{money(data['our_tp'])}</td></tr>
            <tr><td><b>Buy-in price</b></td><td><b>{money(data['buy_in'])}</b></td></tr>
          </tbody>
        </table>
      </div>

      <div class="box">
        <div class="box-title">Consensus / analyst coverage</div>
        <table>
          <thead><tr><th>Casa</th><th>TP</th><th>Rec.</th><th>Rev.</th></tr></thead>
          <tbody>{consensus_rows}</tbody>
        </table>
      </div>

      <div class="box">
        <div class="box-title">Guidance oficial / proximos trimestres</div>
        <ul class="bullets">
          <li><b>Previa 2T26:</b> ROB de R$ 515,0 mi, +9,3% QoQ e +30,6% YoY; colaboradores finais de 7.990 e media ponderada de 8.010.</li>
          <li><span class="tag">Contratos</span> R$ 966,4 mi em novos contratos no 2T26, equivalente a 1,88x a ROB do trimestre.</li>
          <li><span class="tag">Vagas</span> 957 vagas em aberto no fim do 2T26, sinalizando demanda para o trimestre seguinte.</li>
          <li><span class="tag">2026</span> SG&A/ROL ajustado estruturalmente abaixo de 9,5%, versus 11,65% em 2024.</li>
          <li><span class="tag">Backlog</span> 1T26 ja havia trazido R$ 1,42 bi em novos contratos; 2T26 reforca continuidade comercial.</li>
          <li><span class="tag">ROIC</span> roadmap para retorno ao ROIC minimo de 18% via margem, giro de capital, maturacao do backlog e disciplina de SG&A.</li>
        </ul>
      </div>

    </div>
  </div>

  <div class="footer">
    <span>Chief of Staff Digital | PRNR3 single pager | Uso interno, nao e recomendacao de investimento.</span>
    <span>Fontes: RI Priner, MZiQ, Yahoo Finance</span>
  </div>
</section>
</body>
</html>"""


def kpi(label, value, sub="", cls=""):
    return f"<div class='kpi'><div class='kpi-label'>{html.escape(label)}</div><div class='kpi-value {cls}'>{html.escape(value)}</div><div class='kpi-sub'>{html.escape(sub or '')}</div></div>"


def price_kpi(value, day_change):
    arrow_html = ""
    if day_change is not None:
        is_up = day_change >= 0
        color = "#067647" if is_up else "#b42318"
        arrow = "&#9650;" if is_up else "&#9660;"
        arrow_html = (
            f"<span style='color:{color}; font-size:13px; margin-left:6px; "
            f"vertical-align:2px; font-weight:950'>{arrow}</span>"
        )
    sub = signed_pct(day_change) + " hoje"
    return (
        "<div class='kpi'>"
        "<div class='kpi-label'>Price</div>"
        f"<div class='kpi-value'>{html.escape(value)}{arrow_html}</div>"
        f"<div class='kpi-sub'>{html.escape(sub)}</div>"
        "</div>"
    )


def metric(label, value, sub=""):
    sub_html = f"<em>{html.escape(sub)}</em>" if sub else ""
    return f"<div class='metric'><span>{html.escape(label)}</span><b>{html.escape(value)}</b>{sub_html}</div>"


def sparkline_svg(points, width, height, analyst_calls=None, show_legend=True):
    if not points:
        return "<div class='note'>Historico indisponivel.</div>"
    values = [p["close"] for p in points]
    min_v, max_v = min(values), max(values)
    median_v = statistics.median(values)
    pad = 8
    denom = max(max_v - min_v, 0.01)
    coords = []
    for idx, value in enumerate(values):
        x = pad + idx / max(1, len(values) - 1) * (width - 2 * pad)
        y = pad + (max_v - value) / denom * (height - 2 * pad)
        coords.append((x, y))
    path = " ".join(("M" if i == 0 else "L") + f"{x:.1f},{y:.1f}" for i, (x, y) in enumerate(coords))
    start, end = values[0], values[-1]
    color = "#7c3aed"
    area = f"M{coords[0][0]:.1f},{height-pad} " + " ".join(f"L{x:.1f},{y:.1f}" for x, y in coords) + f" L{coords[-1][0]:.1f},{height-pad} Z"
    median_y = pad + (max_v - median_v) / denom * (height - 2 * pad)
    quarter_grid = quarter_grid_svg(points, width, height, pad)
    markers = analyst_call_markers(points, coords, analyst_calls or [])
    legend = analyst_call_legend(width, height) if markers and show_legend else ""
    return f"""
    <svg viewBox="0 0 {width} {height}" width="100%" height="100%" role="img" aria-label="PRNR3 12m chart">
      <rect x="0" y="0" width="{width}" height="{height}" fill="#fff"/>
      {quarter_grid}
      <path d="{area}" fill="{color}" opacity=".10"/>
      <line x1="{pad}" y1="{median_y:.1f}" x2="{width-pad}" y2="{median_y:.1f}" stroke="#94a3b8" stroke-width=".7" stroke-dasharray="4 4" opacity=".70"/>
      <text x="{width-pad}" y="{median_y - 3:.1f}" fill="#667085" font-size="9" text-anchor="end">Median {median_v:.0f}</text>
      <path d="{path}" fill="none" stroke="{color}" stroke-width="1.9"/>
      {markers}
      <text x="{pad}" y="{height-2}" fill="#667085" font-size="10">{points[0]['date'].strftime('%b/%y')}</text>
      <text x="{width-pad}" y="{height-2}" fill="#667085" font-size="10" text-anchor="end">{points[-1]['date'].strftime('%b/%y')}</text>
      <text x="{pad}" y="13" fill="#667085" font-size="10">Low {brl_plain(min_v)}</text>
      <text x="{width-pad}" y="13" fill="#667085" font-size="10" text-anchor="end">High {brl_plain(max_v)}</text>
      {legend}
    </svg>"""


def quarter_grid_svg(points, width, height, pad):
    first_date = points[0]["date"]
    last_date = points[-1]["date"]
    total_days = max((last_date - first_date).days, 1)
    current = next_quarter_end(first_date)
    out = []
    while current <= last_date:
        x = pad + ((current - first_date).days / total_days) * (width - 2 * pad)
        label = f"{str(current.year)[-2:]}Q{((current.month - 1) // 3) + 1}"
        label_x = x
        anchor = "middle"
        if x > width - pad - 24:
            label_x = width - pad - 18
            anchor = "end"
        elif x < pad + 24:
            label_x = pad
            anchor = "start"
        out.append(
            f"<line x1='{x:.1f}' y1='18' x2='{x:.1f}' y2='{height - pad - 11}' "
            "stroke='#d9e2ec' stroke-width='.7'/>"
        )
        out.append(
            f"<text x='{label_x:.1f}' y='{height - 14}' fill='#667085' "
            f"font-size='8.8' font-weight='800' text-anchor='{anchor}'>{label}</text>"
        )
        current = next_quarter_end(current + dt.timedelta(days=1))
    return "\n      ".join(out)


def next_quarter_end(value):
    end = quarter_end_for_date(value)
    if end <= value:
        return quarter_end_for_date(end + dt.timedelta(days=1))
    return end


def quarter_end_for_date(value):
    quarter_end_month = ((value.month - 1) // 3 + 1) * 3
    next_month = dt.date(value.year, quarter_end_month, 1)
    next_month = add_months(next_month, 1)
    return next_month - dt.timedelta(days=1)


def add_months(value, months):
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    return dt.date(year, month, 1)


def analyst_call_legend(width, height):
    return f"""
      <g font-size="9" fill="#667085">
        <circle cx="{width - 128}" cy="10" r="4.2" fill="#16a34a" stroke="#fff" stroke-width="1"/>
        <text x="{width - 121}" y="13">Buy</text>
        <circle cx="{width - 88}" cy="10" r="4.2" fill="#175cd3" stroke="#fff" stroke-width="1"/>
        <text x="{width - 81}" y="13">Hold</text>
        <circle cx="{width - 42}" cy="10" r="4.2" fill="#111827" stroke="#fff" stroke-width="1"/>
        <text x="{width - 35}" y="13">Sell</text>
      </g>"""


def analyst_call_markers(points, coords, analyst_calls):
    if not points or not analyst_calls:
        return ""
    first_date = points[0]["date"]
    last_date = points[-1]["date"]
    out = []
    for call in analyst_calls:
        call_date = parse_iso_date(call.get("call_date"))
        if not call_date or call_date < first_date or call_date > last_date:
            continue
        idx = min(range(len(points)), key=lambda i: abs((points[i]["date"] - call_date).days))
        x, y = coords[idx]
        kind = analyst_call_kind(call.get("rating"))
        fill = analyst_call_color(kind)
        tooltip = (
            f"{call.get('house', 'Analista')} | {kind} ({call.get('rating', 'N/D')}) | "
            f"TP {money(call.get('target'))} | Revisao {call.get('date', 'N/D')}"
        )
        out.append(
            "<circle class='analyst-call' "
            f"cx='{x:.1f}' cy='{y:.1f}' r='5.8' fill='{fill}' stroke='#fff' stroke-width='1.5'>"
            f"<title>{html.escape(tooltip)}</title>"
            "</circle>"
        )
    return "\n      ".join(out)


def analyst_call_kind(rating):
    text = (rating or "").lower()
    if any(token in text for token in ("compra", "buy", "outperform", "overweight")):
        return "Buy"
    if any(token in text for token in ("venda", "sell", "underperform", "underweight")):
        return "Sell"
    return "Hold"


def analyst_call_color(kind):
    return {
        "Buy": "#16a34a",
        "Hold": "#175cd3",
        "Sell": "#111827",
    }.get(kind, "#175cd3")


def parse_iso_date(value):
    if not value:
        return None
    try:
        return dt.date.fromisoformat(str(value))
    except ValueError:
        return None


def bar_line_svg(labels, bars, line, width, height):
    pad_l, pad_r, pad_t, pad_b = 34, 10, 12, 24
    max_bar = max(bars) if bars else 1
    min_line, max_line = min(line), max(line)
    denom_line = max(max_line - min_line, 0.01)
    bar_w = (width - pad_l - pad_r) / max(len(bars), 1) * 0.58
    gap = (width - pad_l - pad_r) / max(len(bars), 1)
    rects = []
    points = []
    for i, (label, value, margin) in enumerate(zip(labels, bars, line)):
        x = pad_l + i * gap + gap * 0.21
        h = value / max_bar * (height - pad_t - pad_b)
        y = height - pad_b - h
        rects.append(f"<rect x='{x:.1f}' y='{y:.1f}' width='{bar_w:.1f}' height='{h:.1f}' fill='#d9e8ff'/>")
        cx = x + bar_w / 2
        cy = pad_t + (max_line - margin) / denom_line * (height - pad_t - pad_b)
        points.append((cx, cy))
        rects.append(f"<text x='{cx:.1f}' y='{height-7}' font-size='8' text-anchor='middle' fill='#667085'>{html.escape(str(label))}</text>")
    path = " ".join(("M" if i == 0 else "L") + f"{x:.1f},{y:.1f}" for i, (x, y) in enumerate(points))
    return f"""
    <svg viewBox="0 0 {width} {height}" width="100%" height="{height}" role="img" aria-label="ROL e margem EBITDA">
      <rect x="0" y="0" width="{width}" height="{height}" fill="#fff"/>
      <text x="2" y="12" font-size="8.5" fill="#667085">R$ mm</text>
      <text x="{width-2}" y="12" font-size="8.5" fill="#667085" text-anchor="end">Margem EBITDA adj.</text>
      {''.join(rects)}
      <path d="{path}" fill="none" stroke="#175cd3" stroke-width="2"/>
      {''.join(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='2.6' fill='#175cd3'/>" for x,y in points)}
    </svg>"""


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def latest_close(points):
    return points[-1]["close"] if points else None


def history_low(points):
    return min((p["close"] for p in points), default=None)


def history_high(points):
    return max((p["close"] for p in points), default=None)


def history_return(points):
    if len(points) < 2 or not points[0]["close"]:
        return None
    return (points[-1]["close"] / points[0]["close"] - 1) * 100


def history_volatility(points):
    closes = [p["close"] for p in points]
    returns = [closes[i] / closes[i - 1] - 1 for i in range(1, len(closes)) if closes[i - 1]]
    if len(returns) < 2:
        return None
    return statistics.stdev(returns) * math.sqrt(252) * 100


def pct_change(base, value):
    if base in (None, 0) or value is None:
        return None
    return (value / base - 1) * 100


def qoq(current, reference):
    if reference in (None, 0) or current is None:
        return None
    return (current / reference - 1) * 100


def signed_pct(value, plus=True):
    if value is None:
        return "N/D"
    prefix = "+" if plus and value >= 0 else ""
    return f"{prefix}{value:.1f}%"


def delta_pp(current, reference):
    if current is None or reference is None:
        return "N/D"
    return f"{(current - reference) * 100:+.1f} p.p."


def money(value):
    if value is None:
        return "N/D"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def brl_plain(value):
    if value is None:
        return "N/D"
    return f"{value:.2f}"


def brl_mm(value):
    if value is None:
        return "N/D"
    return f"R$ {value:,.1f} mm".replace(",", "X").replace(".", ",").replace("X", ".")


def brl_big(value):
    if value is None:
        return "N/D"
    if value >= 1_000_000_000:
        return f"R$ {value / 1_000_000_000:.1f} BI".replace(".", ",")
    if value >= 1_000_000:
        return f"R$ {value / 1_000_000:.0f} mm"
    return money(value)


def normalize_multiples(value):
    return (value or "N/D").replace(" | ", " / ").replace("FWRD", "Fwrd")


def multiple_display(multiples, label):
    for part in (multiples or "").split(" | "):
        if part.startswith(label + " "):
            value = part.replace(label + " ", "")
            return f"{value}x" if value not in ("N/D", "N/A", "0.00") else value
    return "N/D"


def priner_ev_ebitda_display(data):
    financials = data.get("financials", {})
    ltm = financials.get("ltm", {})
    latest = financials.get("latest", {})
    ebitda = ltm.get("ebitda_adj")
    leverage = latest.get("net_debt_ebitda")
    market_cap = data.get("market_cap")
    if ebitda in (None, 0) or leverage is None or market_cap is None:
        return multiple_display(data.get("multiples"), "EV/EBITDA")
    net_debt = leverage * ebitda
    ev = market_cap / 1_000_000 + net_debt
    return f"{ev / ebitda:.2f}x"


def method_display_name(name):
    mapping = {
        "Goldman-style screener": "Goldman Sachs",
        "Morgan Stanley-style DCF": "Morgan Stanley",
        "Bridgewater-style risk": "Bridgewater",
        "JPMorgan-style earnings": "JPMorgan",
        "BlackRock-style portfolio": "BlackRock",
    }
    return mapping.get(name, name.replace("-style", ""))


def valuation_method_row(name, value, max_method, min_method):
    cls = ""
    if max_method is not None and abs(value - max_method) < 0.0001:
        cls = " class='valuation-high'"
    elif min_method is not None and abs(value - min_method) < 0.0001:
        cls = " class='valuation-low'"
    return (
        f"<tr{cls}><td>{html.escape(method_display_name(name))}</td>"
        f"<td>{money(value)}</td></tr>"
    )


if __name__ == "__main__":
    sys.exit(main())
